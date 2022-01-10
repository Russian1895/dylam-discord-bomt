import discord
from help import helpful
from discord.ext import commands, tasks
from discord.utils import get
import openpyxl as xl
from person import person
from discord.ext.commands import Bot
import datetime
import random
import dotenv
import os
import asyncio
env_file = dotenv.dotenv_values()
# MOST OF WHAT I HAVE IMPORTED HERE IS ABSOLUTELY USELESS
import youtube_dl
from openpyxl import Workbook

# VERSION 1.19.5
# special thanks to Andres for teaching me about discord.py rewrite

bday = commands.Bot(command_prefix = ['='], case_insensitive = True, help_command = helpful())
#I don't actually think these vars do anything but I keep them cause they are cute
name1 = ''
name2 = ''
multiple = False

players = {}

# YEAH THESE ARE THE SONG FILES LISTEN I DONT OPTIMIZE CODE VERY WELL OK
choices = ['Akina Nakamori - Anata no Portrait.flac',
'Akina Nakamori - Fragile Afternoon.flac',
'Akina Nakamori - Hot Springs.flac',
'Akina Nakamori - Into the Azure Night.mp3',
'Akina Nakamori - Legend of the Galaxy.flac',
'Akina Nakamori - Mythology.mp3',
'Akina Nakamori - Shojo A.mp3',
'Akina Nakamori - Slow Motion.flac',
'Akina Nakamori - Shipwreck.mp3',
'Akina Nakamori - Jukkai (1984).flac',
'Akina Nakamori - Not a Decoration, Tears.mp3',
'Akina Nakamori - Second Love.flac',
'Akina Nakamori - Southern Wind.flac',
'Akina Nakamori - Shiroi Labyrinth.mp3',
'Akina Nakamori - Horizon.mp3',
'Akina Nakamori - It was Raining.mp3',
'Akina Nakamori - Stripe.mp3',
'Akina Nakamori - Jealous Candle.mp3',
'Anri - SHYNESS BOY.flac',
'Anri - WINDY SUMMER.flac',
'Anri - YOU ARE NOT ALONE.flac',
"Anri - I Can't Stop the Loneliness.flac",
'Chisato Moritaka - The Miha.mp3',
'Hiromi Iwasaki - Duplex.mp3',
'Hiromi Iwasaki - We.mp3',
'Hiroko Yakushimaru - Detective Story.mp3',
'Hiroko Yakushimaru - Main Theme.mp3',
'Hiroko Yakushimaru - Sailor Suit Machine Gun.mp3',
'Hiroko Yakushimaru - Talk About Love.mp3',
'Hitomi Ishikawa - Alone.mp3',
'Hitomi Ishikawa - Heart Communication.mp3',
'Hitomi Ishikawa - Machibuse.mp3',
'Hitomi Ishikawa - Right to Right.mp3',
'Hitomi Ishikawa - Shower.mp3',
'Inoue Nozomi - Lefranc.mp3',
'Kozo Murashita - Hatsukoi.mp3',
'Naoko Kawai - Control.mp3',
'Naughty Boys - Ongaku.mp3',
'Taeko Onuki - 4-00A.M..flac',
'Taeko Onuki - Desolation.flac',
'Taeko Onuki - Law of Nature.flac',
'Taeko Onuki - Metropolis.flac',
'Taeko Onuki - Summer Connection.flac',
'Tatsu Yamashita - Saucy Dog.flac',
'Tatsuro Yamashita - DAYDREAM.flac',
'Tatsuro Yamashita - Ride on Time.flac',
'Tomoyo Harada - Cool.mp3',
'Tomoyo Harada - Rain Planetarium.mp3',
'Wada Kanako - Birthday Minus 1.flac',
'Wada Kanako - Heart.flac',
'Wada Kanako - I Love You.flac',
'Wada Kanako - Like a Salvia Flower.flac',
'Aqua City - Misty night Cruising.mp3',
'Aqua City - Reverside Hotel.mp3',
'CINDY - Believing in Ourselves.flac',
'Yurie Kokubu - Refrain.mp3',
'Yurie Kokubu - WANNA BE WITH YOU.mp3',
'Yurie Kokobu - Just a Joke.mp3',
'Ito Chieri - Merry Christmas.flac',
'Ito Chieri - He Disappeared in the Rain.mp3',
'Junko Yagari - Mr. Blue.flac',
'Yunko Yagami - BAY CITY.flac',
'Kingo Hamada - Dolphin in Town.mp3',
'Kingo Hamada - Midnight Cruisin.mp3',
'Mai Yamane - Tasogare.mp3',
'Mariya Takeuchi - Once Again.flac',
'Mariya Takeuchi - Plastic Love.flac',
'Mariya Takeuchi - September.flac',
'Mariya Takeuchi - End of Love.flac',
'Meiko Nakahara - Fantasy.flac',
'Meiko Nakahara - Friday Magic.flac',
'Meiko Nakahara - Gigolo.flac',
'Meiko Nakahara - Go Away.mp3',
'Meiko Nakahara - Rainy Day.mp3',
'Meiko Nakahara - Ru Ru Russian Roulette.mp3',
'Meiko Nakahara - Scorpion.mp3',
'Meiko Nakahara - Juggler.flac',
'Meiko Nakahara - Sexy dandy.flac',
'Meiko Nakahara - Sleeping Princess.flac',
'Meiko Nakahara - Kiwi Papaya Mango.mp3',
'Miki Fujitani - With Tears, Goodbye.mp3',
'Miki Imai - Kisses in the Rain.mp3',
'Norie Hayashi - Etranze.mp3',
'Omega Tribe - ASPHALT LADY.mp3',
'Junko Ohashi - I Love You So.flac',
'Junko Ohashi - Telephone Number.flac',
'Kyoko Koizumi - Kaze no Magical Mermaid.mp3',
'Riho Makise - Miracle Love.flac',
'Taeko Rei - Love Maiden.mp3',
"Tomoko Aran - I'm In Love.flac",
'Yasuha - Flyby Chinatown.flac',
'Yasuha - Paul Pauly Paul.mp3',
'Yasuha - Short Story.flac',
'Yoshimi Iwasaki - Midnight Fantasy.mp3',
'Yoshimi Iwasaki - Rain.mp3',
"Yoshiko Tanaka - Cote d'Azur.mp3",
'Yoshiko Tanaka - Omoide wa Azayaka ni.mp3',
'Yuki Saito - Axia.flac',
'Yuki Saito - Shiroi Hono.flac',
'Yuki Kato - BLACK JACK.mp3',
'Yukako Hayase - Barefoot Bolero.mp3',
'Yumi Matsutoya - DANG DANG.mp3',
'Yumi Matsutoya - Haruyo.mp3',
"Yumi Matsutoya - Midsummer Night's Dream.mp3",
'Yuming Matsutoya - No Side.mp3',
'Yuming Matsutoya - No-return.flac',
'Yuming Matsutoya - Refrain Something.flac',
'Yuming Matsutoya - Youthful Regret.mp3',
'Miki Matsubara - Stay with me.flac',
'Miki Matsubara - WASH.mp3',
'Momoe Yamaguchi - Akai Shougeki.mp3',
'Momoe Yamaguchi - COSMOS.mp3',
'Momoe Yamaguchi - Iihi Tabidachi.mp3',
'Momoe Yamaguchi - PLAYBACK Part2.mp3',
'Momoe Yamaguchi - Star Shine Dance.mp3',
'Momoe Yamaguchi - Sayonara No Mukougawa.mp3',
'Momoe Yamaguchi - Yokosuka Story.mp3',
'Momoko Kikuchi - Ivory Coast.mp3',
"Momoko Kikuchi - Can't Meet You Anymore.mp3",
'Momoko Kikuchi - Natsuiro Kataomoi.mp3',
'Momoko Kikuchi - ADVENTURE.mp3',
'Momoko Kikuchi - DEJA VU.mp3',
'Momoko Kikuchi - Non Stop the Rain.flac',
'Serbian Apple - Summer Water Rise.mp3',
'Wink - Turn It Into Love.mp3'
]

# THIS BAD BOY HERE CHECKS TO SEE IF THE BOT ACTUALLY WORKS
@bday.event
async def on_ready():
    print('LIBERTY PRIME ONLINE')
    await bday.change_presence(activity = discord.Game('(Prefix is =)'))

# TEST COMMAND TO MAKE SURE THE BOT RESPONDS TO COMMANDS
@bday.command(hidden = False, description = 'shut up')
async def dylan(ctx,*args):
    await ctx.send('shut up')
    pain.start(ctx,False,"dylam bomt")

# MY THANKS
@bday.command(hidden = True, description = 'yes')
async def thanks(ctx,*args):
    await ctx.send('Yeah no problem buddy.')

# MMMM VERY GOOD
@bday.command(hidden = True, description = 'smoothie')
async def smoo(ctx,*args):
    await ctx.send('https://cdn.discordapp.com/attachments/729916916793081956/760370538723672074/smoo.mp4')

# CYCLES NAMES WHEN IT IS SOMEONE'S BIRTHDAY TODAY
@tasks.loop(seconds = 10)
async def pain(ctx, mult, name1, name2 = '', cycle = 0):
    ctx = ctx.author
    await bday.wait_until_ready()
    if mult == False:
        await ctx.guild.me.edit(nick=name1)
    else:
        if cycle == 0:
            await ctx.guild.me.edit(nick=name1)
            cycle = 1
        if cycle == 1:
            await ctx.guild.me.edit(nick=name2)
            cycle = 0

# CHOOSES A RANDOM SONG FROM THE SONGS FOLDER TO GET FUNKY WITH
@bday.command(hidden = False, pass_context = True,description = 'Plays a random song.')
async def song(ctx):
    if ctx.author.voice and ctx.author.voice.channel:
        channel = ctx.author.voice.channel
    else:
        await ctx.send("you aren't connected to a voice channel silly")
        return
    global vc
    try:
        vc=await channel.connect()
    except:
        TimeoutError
    if vc.is_playing():
        vc.stop()

    bum = random.randint(0,len(choices))
    last = choices[bum]
    current = discord.Embed(title = "Now Playing:")
    current.add_field(name = f'Song {bum + 1}/{len(choices)}', value = last)
    await ctx.send(embed = current)
    guild = ctx.message.guild
    voice_client = guild.voice_client
    player = vc.play(discord.FFmpegPCMAudio(f'songs\{last}'), after=lambda e: print('done', e))

# TO PICK SPECIFIC SONGS IN CASE YOU SUCK
@bday.command(hidden = False, pass_context = True,description = 'Pick a list song by #')
async def pick(ctx, num = 1):
    num = num - 1
    if ctx.author.voice and ctx.author.voice.channel:
        channel = ctx.author.voice.channel
    else:
        await ctx.send("you aren't connected to a voice channel dum dum")
        return
    global vc
    try:
        vc=await channel.connect()
    except:
        TimeoutError
    if vc.is_playing():
        vc.stop()
    last = choices[num]
    current = discord.Embed(title = "Now Playing:")
    current.add_field(name = f'Song {num + 1}/{len(choices)}', value = last)
    await ctx.send(embed = current)
    guild = ctx.message.guild
    voice_client = guild.voice_client
    player = vc.play(discord.FFmpegPCMAudio(f'songs\{last}'), after=lambda e: print('done', e))

# GIVES A SHITTY LIST OF ALL THE SONGS IN THE SONG FOLDER
@bday.command(hidden = False, pass_context = True,description = 'List for =song.')
async def slist(ctx):
    sorry = choices.copy()
    first = sorry[:len(sorry)//3]
    second = sorry[len(sorry)//3: - len(sorry)//3]
    temp = sorry[len(sorry)//3:]
    third = temp[len(temp)//2:]
    for each in first:
        first[first.index(each)] = f'{first.index(each) + 1}: {each}'
    for each in second:
        second[second.index(each)] = f'{second.index(each) + 42}: {each}'
    for each in third:
        third[third.index(each)] = f'{third.index(each) + 84}: {each}'

    dog = discord.Embed(title = "List o' songs(1/3)")
    dog.add_field(name = f'There are currently {len(sorry)} songs in the =song list.',value="Please tell me if any of this doesn't play properly")
    dog.set_footer(text='\n'.join(first))
    await ctx.send(embed = dog)
    jog = discord.Embed(title = "List o' songs(2/3)")
    jog.add_field(name = f"please don't feel overwhelmed",value="this folder is like 1.95 gigabytes")
    jog.set_footer(text='\n'.join(second))
    await ctx.send(embed = jog)
    bog = discord.Embed(title = "List o' songs(3/3)")
    bog.add_field(name = f"can i get a",value="big mac")
    bog.set_footer(text='\n'.join(third))
    await ctx.send(embed = bog)

# WOW THATS A LARGE LIST
@bday.command(hidden = False, description = "List for =next")
async def blist(ctx,*args):
    guys = xl.load_workbook('homies.xlsx')
    yeet = guys['Sheet1']
    thing = yeet.cell(1, 1)
    people = []
    f = []
    l = []
    date = []

    for row in range(yeet.max_row):
        first = yeet.cell(row+1, 1)
        last = yeet.cell(row+1, 2)
        mon = yeet.cell(row+1, 3)
        day = yeet.cell(row+1, 4)
        yea = yeet.cell(row+1, 5)
        lamer = person(first.value,mon.value,day.value,yea.value,last.value)
        f.append(first.value)
        if last.value == None:
            l.append(' ')
        else:
            l.append(last.value)
        date.append(f'{str(mon.value)}/{str(day.value)}/{str(yea.value)}')
        people.append(lamer)

    pog = discord.Embed(title = "List o' Birthdays")
    pog.add_field(name = 'First Name',value = '\n'.join(f))
    pog.add_field(name = 'Last Name', value = '\n'.join(l))
    pog.add_field(name = 'M/D/Y', value = '\n'.join(date))
    pog.set_footer(text = f'There are currently {len(people)} people in the =next list.')
    await ctx.send(embed = pog)

# SECRET COMMANDS FOR SECRET PEOPLE
@bday.command(hidden = True, pass_context = True,description = 'metal gear')
async def snake(ctx):
    if ctx.author.voice and ctx.author.voice.channel:
        channel = ctx.author.voice.channel
    else:
        await ctx.send("you aren't connected to a voice channel silly")
        return
    global vc
    try:
        vc=await channel.connect()
    except:
        TimeoutError
    if vc.is_playing():
        vc.stop()

    current = discord.Embed(title = "Now (secretly) Playing:")
    current.add_field(name = 'Snake Eater.mp3', value = 'ladder moment')
    await ctx.send(embed = current)
    guild = ctx.message.guild
    voice_client = guild.voice_client
    player = vc.play(discord.FFmpegPCMAudio('songs\Snake Eater.mp3'), after=lambda e: print('done', e))

@bday.command(hidden = True, pass_context = True,description = 'I WAS A FOOL')
async def anime(ctx):
    if ctx.author.voice and ctx.author.voice.channel:
        channel = ctx.author.voice.channel
    else:
        await ctx.send("baka")
        return
    global vc
    try:
        vc=await channel.connect()
    except:
        TimeoutError
    if vc.is_playing():
        vc.stop()

    current = discord.Embed(title = "Now (secretly) Playing:")
    current.add_field(name = 'bakamitai.mp3', value = 'taxi cab version i think')
    await ctx.send(embed = current)
    guild = ctx.message.guild
    voice_client = guild.voice_client
    player = vc.play(discord.FFmpegPCMAudio('songs\sakamitai.mp3'), after=lambda e: print('done', e))

@bday.command(hidden = True, pass_context = True,description = 'oh really?')
async def best(ctx):
    if ctx.author.voice and ctx.author.voice.channel:
        channel = ctx.author.voice.channel
    else:
        await ctx.send("YOU ARE NOT IN VC SCUM")
        return
    global vc
    try:
        vc=await channel.connect()
    except:
        TimeoutError
    if vc.is_playing():
        vc.stop()
    two = ['Yi Jian Mei(2010).mp3','Yi Jian Mei(1983).mp3']
    bum = random.randint(0,1)
    last = two[bum]
    current = discord.Embed(title = "Now (secretly) Playing:")
    current.add_field(name = last, value = 'oh man could it be the old one or the new one')
    await ctx.send(embed = current)
    guild = ctx.message.guild
    voice_client = guild.voice_client
    player = vc.play(discord.FFmpegPCMAudio(f'songs\{last}'), after=lambda e: print('done', e))

@bday.command(hidden = False, pass_context = True,description = 'from the fantasticks')
async def fant(ctx):
    if ctx.author.voice and ctx.author.voice.channel:
        channel = ctx.author.voice.channel
    else:
        await ctx.send("too much moisture")
        return
    global vc
    try:
        vc=await channel.connect()
    except:
        TimeoutError
    if vc.is_playing():
        vc.stop()
    two = ['monologue.mp3','muchmore(1960).mp3','muchmorenoteblocks.mp3','overture(1960).mp3','overturenoteblock.mp3']
    bum = random.randint(0,4)
    last = two[bum]
    current = discord.Embed(title = "Now Playing:")
    current.add_field(name = last, value = 'Too vibrant for a name...')
    await ctx.send(embed = current)
    guild = ctx.message.guild
    voice_client = guild.voice_client
    player = vc.play(discord.FFmpegPCMAudio(f'songs\{last}'), after=lambda e: print('done', e))

# SECRET COMMAND FOR SECRET SECRET PEOPLE
@bday.command(hidden = True, pass_context = True,description = 'REPENT SINNER')
async def hellfire(ctx):
    if ctx.author.voice and ctx.author.voice.channel:
        channel = ctx.author.voice.channel
    else:
        await ctx.send("you aren't connected to a voice channel *sinner*")
        return
    global vc
    try:
        vc=await channel.connect()
    except:
        TimeoutError
    if vc.is_playing():
        vc.stop()

    current = discord.Embed(title = "Now (secretly) Playing:")
    current.add_field(name = 'hellfire.mp3', value = "swear to notch if they make this into a live action im gonna scream")
    await ctx.send(embed = current)
    guild = ctx.message.guild
    voice_client = guild.voice_client
    player = vc.play(discord.FFmpegPCMAudio('songs\hellfire.mp3'), after=lambda e: print('done', e))

@bday.command(hidden = True, pass_context = True,description = 'pure evil')
async def evil(ctx):
    if ctx.author.voice and ctx.author.voice.channel:
        channel = ctx.author.voice.channel
    else:
        await ctx.send("i hate you")
        return
    global vc
    try:
        vc=await channel.connect()
    except:
        TimeoutError
    if vc.is_playing():
        vc.stop()

    current = discord.Embed(title = "Now (secretly) Playing:")
    current.add_field(name = 'dance.mp3', value = "You're worst nightmare.")
    await ctx.send(embed = current)
    guild = ctx.message.guild
    voice_client = guild.voice_client
    player = vc.play(discord.FFmpegPCMAudio('songs\dance.mp3'), after=lambda e: print('done', e))

# STOPS WHATEVER THE BOT WAS PLAYING AND DISCONNECTS IT FROM THE VC
@bday.command(hidden = False, pass_context = True,description = 'Silence da bot.')
async def stop(ctx):
    if ctx.author.voice and ctx.author.voice.channel:
        channel = ctx.author.voice.channel
    else:
        await ctx.send("the bot isn't in a voice channel dummy")
        return
    global vc
    try:
        await ctx.send("*rude*")
        vc.stop()
        await vc.disconnect()
    except:
        TimeoutError

@bday.command(hidden = True, pass_context = True,description = 'Silence da bot.')
async def leave(ctx):
    if ctx.author.voice and ctx.author.voice.channel:
        channel = ctx.author.voice.channel
    else:
        await ctx.send("the bot isn't in a voice channel dummy")
        return
    global vc
    try:
        await ctx.send("*rude*")
        vc.stop()
        await vc.disconnect()
    except:
        TimeoutError

# ANOTHER GLORIOUS SECRET COMMAND
@bday.command(hidden = True, description = 'secret')
async def die(ctx,*args):
    await ctx.send("I don't think so.")
    await ctx.send("https://youtu.be/MvsyxAHcGmo?t=38")

# MURDER MOMENT
@bday.command(hidden = True, description = 'secret death')
async def kill(ctx,name):
    temp = name.lower()
    if temp == 'dylan' or temp == 'dylam' or temp == 'haworth' or temp == 'dylanhaworth' or temp == 'dylantheidiot' or temp == 'dylambot' or temp == 'haworth,dylan' or temp == 'Dylan_Haworth' or temp == 'dylan_haworth':
        await ctx.send(f'ah')
        await asyncio.sleep(0.5)
        await ctx.send(f'ah')
        await asyncio.sleep(0.5)
        await ctx.send(f'ah')
        await asyncio.sleep(1)
        await ctx.send(f'*not so fast*')
    elif temp == 'jack':
        await ctx.send(f'Sorry, Jack is not spared.')
    else:
        chance = random.randint(0,1)
        if chance == 0:
            await ctx.send(f'{name} dies')
        else:
            await ctx.send(f'{name} had divine shield and lived!')

# SECRET COMMAND THAT MAY OR MAY NOT ACTUALLY WORK
# ADD SWITCH CASE DICTIONARIES
# switch = 0
@bday.command(hidden = True, description = 'very funny thanks')
async def funny(ctx,*args):
#     if switch == 0:
    await ctx.send("https://cdn.discordapp.com/attachments/729916916793081956/752758446767472640/video0-1.mp4")
#         switch = 1
#     elif switch == 1:
#         await ctx.send("https://cdn.discordapp.com/attachments/671538516005748750/761047936649134090/kuh-tanzt.mp4")
#         switch = 2
#     elif switch == 2:
#         await ctx.send("https://cdn.discordapp.com/attachments/671538516005748750/761047967784108032/i_wonder_what_shes_listening_to.mp4")
#         switch = 3
#     elif switch == 3:
#         await ctx.send("https://cdn.discordapp.com/attachments/671538516005748750/761047117950746674/weezer.mp4")
#         switch = 0
#     else:
#         await ctx.send("https://media.discordapp.net/attachments/671538516005748750/761048998009897000/Capture.PNG")
#         switch = 0

# TELLS A RANDOM JOKE (sort of)
@bday.command(hidden = False, description = "Tells a joke.")
async def joke(ctx,*args):
    await ctx.send('*you*')

# @bday.command(hidden = True, description = "CONVERT PLS")
# async def fuck(ctx,*args):
#
#     first = xl.load_workbook('first.xlsx')
#     # sheet = first.get_sheet_by_name('Sheet1')
#     beg = first['Sheet1']
#     id = []
#     name = []
#     address = []
#     apt = []
#     city = []
#
#     for row in range(beg.max_row):
#         if row%6 == 0:
#             if row != 0:
#                 id.append(beg.cell(row,1))
#                 name.append(beg.cell(row-4,1))
#                 address.append(beg.cell(row-3,1))
#                 apt.append(beg.cell(row-2,1))
#                 city.append(beg.cell(row-1,1))
#     for row in range(beg.max_row):
#         if row%6 == 0:
#             if row != 0:
#                 id.append(beg.cell(row,2))
#                 name.append(beg.cell(row-4,2))
#                 address.append(beg.cell(row-3,2))
#                 apt.append(beg.cell(row-2,2))
#                 city.append(beg.cell(row-1,2))
#     for row in range(beg.max_row):
#         if row%6 == 0:
#             if row != 0:
#                 id.append(beg.cell(row,3))
#                 name.append(beg.cell(row-4,3))
#                 address.append(beg.cell(row-3,3))
#                 apt.append(beg.cell(row-2,3))
#                 city.append(beg.cell(row-1,3))
#     book = Workbook()
#     sheet = book.active
#     sheet['A1'] = 'ID'
#     sheet['B1'] = 'NAME'
#     sheet['C1'] = 'ADDRESS'
#     sheet['D1'] = 'APT'
#     sheet['E1'] = 'CITY'
#     for each in range(len(id)):
#         # print(id[each].value)
#         sheet[f'A{each + 2}'] = id[each].value
#         sheet[f'B{each + 2}'] = name[each].value
#         sheet[f'C{each + 2}'] = address[each].value
#         sheet[f'D{each + 2}'] = apt[each].value
#         sheet[f'E{each + 2}'] = city[each].value
#     book.save("product.xlsx")


# HAHA ITS ME BUT VERY QUIET
@bday.command(hidden = True, description = "Tells a joke.")
async def talk(ctx,*args):
    if ctx.author.voice and ctx.author.voice.channel:
        channel = ctx.author.voice.channel
    else:
        await ctx.send("Sorry, I can't seem to connect to your channel ):")
        return
    global vc
    try:
        vc=await channel.connect()
    except:
        TimeoutError
    if vc.is_playing():
        vc.stop()
    guild = ctx.message.guild
    voice_client = guild.voice_client
    player = vc.play(discord.FFmpegPCMAudio('songs\me.m4a'), after=lambda e: print('done', e))

# LITERALLY JUST BDAY FROM AP COMPUTER SCIENCE
@bday.command(hidden = False, description = "Who's birthday is next?")
async def next(ctx,*args):

    await ctx.send('THIS COMMAND BROKE AHHHHHHHHHHHHHH')

    # homies = xl.load_workbook('homies.xlsx')
    # sheet = homies['Sheet1']
    # thing = sheet.cell(1, 1)
    # people = []
    # parallel = []
    # today = datetime.date.today()
    #
    # for row in range(sheet.max_row):
    #     first = sheet.cell(row+1, 1)
    #     last = sheet.cell(row+1, 2)
    #     mon = sheet.cell(row+1, 3)
    #     day = sheet.cell(row+1, 4)
    #     yea = sheet.cell(row+1, 5)
    #     gamer = person(first.value,mon.value,day.value,yea.value,last.value)
    #     people.append(gamer)
    #
    # for filter in range(len(people)):
    #     if people[filter].bdate.month == today.month:
    #         parallel.append(people[filter])
    #
    # if len(parallel) == 0:
    #     await ctx.send('There are no birthdays this month...')
    #     return
    # else:
    #     closest = []
    #     othertemp = 0
    #     for h in range(len(parallel)):
    #         if (parallel[h].bdate.day > today.day) or (parallel[h].bdate.day == today.day):
    #             othertemp = othertemp + 1
    #     if othertemp == 0:
    #         await ctx.send('There are no more birthdays this month!')
    #         return
    #
    # chumpy = 0
    # poopdex = 0
    # peedex = 0
    #
    # for b in range(len(parallel)):
    #     if parallel[b].bdate.day == today.day:
    #         if chumpy == 0:
    #             chumpy = chumpy + 1
    #             poopdex = b
    #         if chumpy > 0:
    #             chumpy = chumpy + 1
    #             peedex = b
    #     if parallel[b].bdate.day > today.day:
    #         closest.append(parallel[b].bdate.day - today.day)
    #
    # bing = 'thing'
    # ding = 'wing'
    # ing = 'hell'
    #
    # if parallel[poopdex].lname == '':
    #     bing = ''
    # else:
    #     bing = ' '
    # if parallel[peedex].lname == '':
    #     ding = ''
    # else:
    #     ding = ' '
    #
    # if chumpy == 2:
    #     await ctx.send(f"My boy {parallel[poopdex].fname}{bing}{parallel[poopdex].lname}'s birthday is today!")
    #     multiple = False
    #     name1 = (f"Birthday!: {parallel[poopdex].fname}{bing}{parallel[poopdex].lname}")
    #     pain.start(ctx,False,name1)
    #     return
    #
    # if chumpy > 1:
    #     await ctx.send(f"My boys {parallel[poopdex].fname}{bing}{parallel[poopdex].lname} and {parallel[peedex].fname}{ding}{parallel[peedex].lname}'s birthdays are today!")
    #     name1 = (f"Birthday!: {parallel[poopdex].fname}{bing}{parallel[poopdex].lname}")
    #     name2 = (f"Birthday!: {parallel[peedex].fname}{ding}{parallel[peedex].lname}")
    #     multiple = True
    #     pain.start(ctx,True,name1,name2)
    #     return
    #
    # min = 34
    # index = 0
    # twindex = 0
    # twins = False
    #
    # for c in range(len(closest)):
    #     if closest[c] < min:
    #         min = closest[c]
    #         index = c
    # for f in range(len(closest)):
    #     if closest[f] == min and f != index:
    #         twindex = f
    #         twins = True
    #
    # temp = today.year - parallel[index].bdate.year
    # hemp = today.year - parallel[twindex].bdate.year
    #
    # if parallel[twindex].lname == '':
    #     ing = ''
    # else:
    #     ing = ' '
    #
    # if (min == 1 and twins == False) or (min > 1 and twins == False):
    #     await ctx.send(f'My boy {parallel[index].fname}{bing}{parallel[index].lname} is turning {temp} in {min} day(s)!')
    #     pain.start(ctx)
    #     return
    # if (min == 1 and twins == True) or (min > 1 and twins == True):
    #     await ctx.send(f"My boys {parallel[index].fname}{bing}{parallel[index].lname} and {parallel[twindex].fname}{ing}{parallel[twindex].lname} are turning {temp} and {hemp} in {min} day(s)!")
    #     pain.start(ctx)
    #     return

# GENERAL ERROR OUTPUT IN CASE YOU SUCK
@bday.event
async def on_command_error(ctx,error):
    print('YO MR. WHITE THERE WAS AN ERROR')
    await ctx.send('Jesse what are you talking about')
    await ctx.send('https://i.insider.com/5dade9bc045a3139e8686c33?width=1136&format=jpeg')

bday.run(env_file['TOKEN'])
